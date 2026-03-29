"""
Excel questionnaire parser — LLM-based row classification.

Pipeline:
  1. openpyxl extracts per-cell metadata (value, bold, bg color, merged range, data_validation)
  2. Rows serialised to compact text and sent to LLM in batches of BATCH_SIZE
  3. LLM returns per-row classification: QUESTION / SECTION_HEADER / INSTRUCTION / OTHER
  4. Section hierarchy tracked; questions returned with section context + answer cell ref
"""

import json
import logging
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openai import AsyncOpenAI

from config import settings

logger = logging.getLogger(__name__)

BATCH_SIZE = 40


# ---------------------------------------------------------------------------
# Step 1: Extract row metadata from Excel
# ---------------------------------------------------------------------------

def _score_sheet(ws) -> int:
    """Heuristic: count rows that look like questions/data (not pure instructions)."""
    score = 0
    for row in ws.iter_rows(values_only=True):
        vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if not vals:
            continue
        text = " ".join(vals)
        if any(c in text for c in "?？") or len(text) > 20:
            score += 1
    return score


def _extract_excel_rows(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)

    # Pick the sheet most likely to contain questions
    best_ws = wb.active
    best_score = _score_sheet(best_ws)
    for name in wb.sheetnames:
        ws_candidate = wb[name]
        s = _score_sheet(ws_candidate)
        if s > best_score:
            best_score = s
            best_ws = ws_candidate
    ws = best_ws

    max_col = ws.max_column or 1

    # Build lookup: (row, col) -> merge flags
    merge_lookup: dict[tuple, dict] = {}
    for mr in ws.merged_cells.ranges:
        full_width = (mr.min_col == 1 and mr.max_col >= max(max_col - 1, 1))
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_lookup[(r, c)] = {
                    "top_left": (r == mr.min_row and c == mr.min_col),
                    "full_width": full_width,
                }

    # Build set of cell coordinates that have data validation (dropdowns etc.)
    validated_coords: set[str] = set()
    try:
        for dv in ws.data_validations.dataValidation:
            if dv.type:
                for cell_range in dv.sqref.ranges:
                    for r in range(cell_range.min_row, cell_range.max_row + 1):
                        for c in range(cell_range.min_col, cell_range.max_col + 1):
                            validated_coords.add(f"{get_column_letter(c)}{r}")
    except Exception:
        pass

    rows: list[dict] = []
    for excel_row in ws.iter_rows():
        row_idx = excel_row[0].row
        cells_data: list[dict] = []

        for cell in excel_row:
            val = "" if cell.value is None else str(cell.value).strip()

            bold = False
            try:
                bold = bool(cell.font and cell.font.bold)
            except Exception:
                pass

            has_bg = False
            try:
                if cell.fill and cell.fill.fill_type not in (None, "none"):
                    rgb = cell.fill.fgColor.rgb
                    has_bg = bool(rgb and rgb not in ("00000000", "FFFFFFFF"))
            except Exception:
                pass

            has_validation = cell.coordinate in validated_coords

            mi = merge_lookup.get((cell.row, cell.column), {})
            # Include cells with content, merge top-left, or dropdown (even if empty)
            if val or mi.get("top_left") or has_validation:
                cells_data.append({
                    "col": cell.column_letter,
                    "val": val,
                    "bold": bold,
                    "bg": has_bg,
                    "full_merge": mi.get("full_width", False),
                    "dropdown": has_validation,
                })

        if cells_data:
            rows.append({"row": row_idx, "cells": cells_data})

    wb.close()
    return rows


# ---------------------------------------------------------------------------
# Step 2: Serialise rows for LLM
# ---------------------------------------------------------------------------

def _serialize_rows(rows: list[dict]) -> str:
    lines: list[str] = []
    for r in rows:
        parts: list[str] = []
        for c in r["cells"]:
            flags = ""
            if c["bold"]:
                flags += " bold"
            if c["bg"]:
                flags += " shaded"
            if c["full_merge"]:
                flags += " fullmerge"
            if c["dropdown"]:
                flags += " dropdown"
            val_preview = c["val"][:100].replace("\n", " ")
            parts.append(f"[{c['col']}:{val_preview}{flags}]")
        lines.append(f"R{r['row']}: " + "  ".join(parts))
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Step 3: LLM batch classification
# ---------------------------------------------------------------------------

_SYSTEM = """You are an expert at parsing compliance and security questionnaire spreadsheets.
You will receive rows from an Excel file with formatting hints:
  bold = bold text  |  shaded = colored background  |  fullmerge = merged across all columns  |  dropdown = has answer dropdown

Classify each row. Types:
  SECTION_HEADER  - a category/section title (typically fullmerge+bold or shaded, not a question)
  QUESTION        - a row that the vendor/respondent must fill in an answer for
  INSTRUCTION     - pure explanatory notes, submission instructions, or contact info
  OTHER           - column headers, blank rows, metadata rows with no response needed

QUESTION includes ALL of these patterns:
  - Direct questions ("Do you have...?", "How often...?")
  - Statement-style requirements with an adjacent answer cell ("Data retention period for customer records")
  - Compliance checklist items where Yes/No/N/A must be selected (especially rows with +dropdown)
  - Data request fields where a value must be filled in ("Company legal name", "Date of last audit")
  - Any row where a human is clearly expected to write or select a response

For QUESTION rows output:
- question_col: the column letter containing the actual question/requirement text. If a row has a letter/number index in col A and the real question in col C, return C. If there is only one text column, return null.
- answer_col: the column letter where the answer goes (empty/dropdown/placeholder column).

Rules:
- Full-width merged rows (+fullmerge) spanning the sheet are almost always SECTION_HEADER.
- A row whose adjacent column has +dropdown is almost certainly a QUESTION.
- Rows like "Please complete all fields below" or "Submit by email to..." are INSTRUCTION.
- Column header rows (e.g. "Question | Response | Required") are OTHER.
- When in doubt between QUESTION and INSTRUCTION, prefer QUESTION if there is an empty adjacent cell."""

_USER_TMPL = """Rows:
{rows}

Return ONLY a JSON object: {{"items": [{{"row": <int>, "type": "<TYPE>", "question_col": "<col letter or null>", "answer_col": "<col letter or null>"}}]}}
- question_col: the column that contains the question/requirement text (null if obvious — single text column)
- answer_col: the column where the answer goes (empty/dropdown/placeholder column)
Include every non-blank row. Omit completely blank rows."""


async def _classify_batch(rows: list[dict], client: AsyncOpenAI) -> list[dict]:
    serialized = _serialize_rows(rows)
    try:
        resp = await client.chat.completions.create(
            model=settings.openai_model,
            messages=[
                {"role": "system", "content": _SYSTEM},
                {"role": "user", "content": _USER_TMPL.format(rows=serialized)},
            ],
            temperature=0,
            response_format={"type": "json_object"},
        )
        content = resp.choices[0].message.content
        try:
            data = json.loads(content)
        except json.JSONDecodeError:
            # LLM occasionally returns two concatenated JSON objects; take the first
            data, _ = json.JSONDecoder().raw_decode(content)
        items = data.get("items", []) if isinstance(data, dict) else data
        return items if isinstance(items, list) else []
    except Exception as e:
        logger.error("_classify_batch failed (rows %d-%d): %r", rows[0]["row"], rows[-1]["row"], e)
        return []


# ---------------------------------------------------------------------------
# Step 4: Build question list from classified rows
# ---------------------------------------------------------------------------

def _build_questions(rows: list[dict], classified: list[dict]) -> list[dict]:
    cls_by_num = {
        item["row"]: item
        for item in classified
        if isinstance(item, dict) and "row" in item
    }

    current_section = ""
    questions: list[dict] = []
    seq = 0

    for row in rows:
        rnum = row["row"]
        cls = cls_by_num.get(rnum)
        if not cls:
            continue

        rtype = cls.get("type", "OTHER")

        if rtype == "SECTION_HEADER":
            text = next((c["val"] for c in row["cells"] if c["val"]), "")
            current_section = text

        elif rtype == "QUESTION":
            question_col = cls.get("question_col")
            if question_col and re.match(r"^[A-Z]+$", str(question_col)):
                # Use only the designated question column
                text = next((c["val"] for c in row["cells"] if c["col"] == question_col), "").strip()
            else:
                # Fallback: use the longest non-empty cell value
                candidates = [c["val"] for c in row["cells"] if c["val"]]
                text = max(candidates, key=len, default="").strip()
            if not text:
                continue

            answer_col = cls.get("answer_col")
            answer_cell = (
                f"{answer_col}{rnum}"
                if answer_col and re.match(r"^[A-Z]+$", str(answer_col))
                else None
            )

            questions.append({
                "seq": seq,
                "question_text": text,
                "section": current_section or None,
                "answer_cell": answer_cell,
            })
            seq += 1

    return questions


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

async def parse_excel_questionnaire_llm(path: str) -> list[dict]:
    """
    Parse an Excel questionnaire using LLM row classification.
    Returns list[dict]: {seq, question_text, section, answer_cell}
    """
    rows = _extract_excel_rows(path)
    if not rows:
        return []

    client = AsyncOpenAI(api_key=settings.openai_api_key, base_url=settings.openai_base_url)

    all_classified: list[dict] = []
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i:i + BATCH_SIZE]
        classified = await _classify_batch(batch, client)
        all_classified.extend(classified)

    return _build_questions(rows, all_classified)


async def parse_questionnaire_file_llm(path: str, filename: str) -> list[dict]:
    """
    Dispatcher for different file types.
    Returns list[dict]: {seq, question_text, section, answer_cell}
    """
    suffix = Path(filename).suffix.lower()

    if suffix in (".xlsx", ".xls"):
        return await parse_excel_questionnaire_llm(path)

    elif suffix == ".txt":
        lines = Path(path).read_text(encoding="utf-8").splitlines()
        return [
            {"seq": i, "question_text": line.strip(), "section": None, "answer_cell": None}
            for i, line in enumerate(lines)
            if line.strip()
        ]

    else:
        raise ValueError(f"Unsupported questionnaire format: {suffix}")
