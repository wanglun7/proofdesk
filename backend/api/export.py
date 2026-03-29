import io
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from api.deps import require_workspace_member
from api.scoping import get_questionnaire_for_workspace
from auth_utils import AuthContext
from database import get_db
from models import Questionnaire, Question, Answer

router = APIRouter()

QUESTIONNAIRE_FILES_DIR = Path(__file__).parent.parent / "questionnaire_files"


def _ensure_questionnaire_exportable(rows: list[tuple[Question, Answer]]) -> None:
    if not rows:
        raise HTTPException(404, "No answers found")
    if any(answer.status != "approved" for _, answer in rows):
        raise HTTPException(409, "All answers must be approved before export")


def _build_summary_workbook(
    questionnaire: Questionnaire,
    rows: list[tuple[Question, Answer]],
) -> tuple[io.BytesIO, str]:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Answers"

    headers = ["#", "Question", "Answer", "Confidence", "Status", "Sources"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    for row_idx, (question, answer) in enumerate(rows, 2):
        final_answer = answer.human_edit or answer.draft or ""
        sources = "; ".join(
            f"{c['source']} p.{c['page']}"
            for c in (answer.citations or [])
            if isinstance(c, dict) and c.get("source") and c.get("page") is not None
        )
        ws.cell(row=row_idx, column=1, value=question.seq + 1)
        ws.cell(row=row_idx, column=2, value=question.question_text)
        cell = ws.cell(row=row_idx, column=3, value=final_answer)
        cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=row_idx, column=4, value=round(answer.confidence or 0, 2))
        ws.cell(row=row_idx, column=5, value=answer.status)
        ws.cell(row=row_idx, column=6, value=sources)

    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_filename = questionnaire.filename.rsplit(".", 1)[0] + "_answered.xlsx"
    return buf, safe_filename


async def _load_export_rows(
    qid: str,
    db: AsyncSession,
    workspace_id,
) -> tuple[Questionnaire, list[tuple[Question, Answer]]]:
    questionnaire = await get_questionnaire_for_workspace(db, qid, workspace_id)
    rows_result = await db.execute(
        select(Question, Answer)
        .join(Answer, Answer.question_id == Question.id)
        .where(Question.questionnaire_id == qid)
        .order_by(Question.seq)
    )
    rows = rows_result.all()
    _ensure_questionnaire_exportable(rows)
    return questionnaire, rows


@router.get("/questionnaire/{qid}/export")
async def export_questionnaire(
    qid: str,
    db: AsyncSession = Depends(get_db),
    auth: AuthContext = Depends(require_workspace_member),
):
    q, rows = await _load_export_rows(qid, db, auth.workspace_id)
    buf, safe_filename = _build_summary_workbook(q, rows)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
    )


@router.get("/questionnaire/{qid}/export-filled")
async def export_questionnaire_filled(
    qid: str,
    db: AsyncSession = Depends(get_db),
    auth: AuthContext = Depends(require_workspace_member),
):
    q, rows = await _load_export_rows(qid, db, auth.workspace_id)

    stored_path = QUESTIONNAIRE_FILES_DIR / f"{qid}.xlsx"
    has_answer_cells = any(question.answer_cell for question, _ in rows)

    if not stored_path.exists() or not has_answer_cells:
        buf, safe_filename = _build_summary_workbook(q, rows)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'},
        )

    wb = openpyxl.load_workbook(stored_path)

    # Pick the same sheet the parser chose (highest question/data row count)
    def _score_sheet(ws) -> int:
        score = 0
        for row in ws.iter_rows(values_only=True):
            vals = [str(v).strip() for v in row if v is not None and str(v).strip()]
            if not vals:
                continue
            text = " ".join(vals)
            if any(c in text for c in "?？") or len(text) > 20:
                score += 1
        return score

    best_ws = wb.active
    best_score = _score_sheet(best_ws)
    for name in wb.sheetnames:
        candidate = wb[name]
        s = _score_sheet(candidate)
        if s > best_score:
            best_score = s
            best_ws = candidate
    ws = best_ws

    for question, answer in rows:
        if not question.answer_cell:
            continue
        final_answer = answer.human_edit or answer.draft or ""
        if final_answer:
            ws[question.answer_cell] = final_answer

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    stem = q.filename.rsplit(".", 1)[0]
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{stem}_filled.xlsx"'},
    )
