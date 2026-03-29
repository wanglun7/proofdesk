from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
QUESTIONNAIRE_PATH = ROOT / "test_kb" / "test-questionnaire.xlsx"


def test_test_questionnaire_fixture_has_expected_shape(hr_manual_gold):
    wb = openpyxl.load_workbook(QUESTIONNAIRE_PATH, data_only=True)
    ws = wb.active

    assert ws["A1"].value == "Question"
    assert ws["B1"].value == "Answer"
    assert ws["C1"].value == "Notes"

    for index, item in enumerate(hr_manual_gold, start=2):
        assert ws[f"A{index}"].value == item["question_text"]
        assert item["answer_cell"] == f"B{index}"
        assert ws[item["answer_cell"]].value is None

    assert ws["A12"].value is None
    wb.close()
